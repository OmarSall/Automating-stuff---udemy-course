import openpyxl, os

os.chdir('E:\\Python Udemy\\Automate boting stuff with Python programming course\\excel')


wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1']  == 42

wb.save('new_example.xlsx')


############################

# openpyxl.load_workbook('somefile.xlsx')

sheet2 = wb.create_sheet()
wb.get_sheet_names()

sheet2.title = 'My New Sheet Name'
wb.save('example2.xlsx')

wb.create_sheet(index = 0, title = 'My other Sheet')
wb.save('example3.xlsx')
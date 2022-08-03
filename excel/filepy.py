import openpyxl, os

os.chdir('E:\\Python Udemy\\Automate boting stuff with Python programming course\\excel')

workbook = openpyxl.load_workbook('example.xlsx')

# print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')

print(workbook.get_sheet_names())

cell = sheet['A1']
cell.value

sheet.cell(row = 1, column = 2)

for i in range(1,8):
    print(i, sheet.cell(row = i, column = 2).value)